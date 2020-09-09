# import cv2
import numpy as np
# import tensorflow as tf
import random
import struct
import bitarray
import csv
import openpyxl

if __name__ == "__main__":
    fp_gen_num = 10
    fp_mat = []
    for idx in range(fp_gen_num):
        x_f32 = random.uniform(-10, 20)
        x_f16 = np.float16(x_f32)
        fp_mat.append(x_f16)

    print(fp_mat[0])
    print(bin(np.float16(0).view('H'))[2:].zfill(16))
    fp_a = np.array(fp_mat, dtype='float16')
    fp_b = fp_mat[1:]
    fp_b.append(fp_a[0])
    fp_b = np.array(fp_b, dtype='float16')

    print(fp_a, fp_a.dtype)
    print(fp_b, fp_b.dtype)
    fp_c = np.array(fp_a) * np.array(fp_b)
    fp_c = np.float16(fp_c)
    print(fp_c, fp_c.dtype)

    test_mat = [fp_a,  fp_b , fp_c]
    test_mat_T = np.transpose(test_mat)
    print(test_mat_T)
    print(test_mat_T[1,1])

    wb = openpyxl.Workbook()
    sheet = wb.active
    f = open('test3.txt','w')

    for idy in range(fp_gen_num):
        for idx in range(3):
            sheet.cell(row=idy + 1, column=2*idx+1).value = test_mat_T[idy, idx]
            sheet.cell(row=idy+1, column=2*idx+2).value = bin(np.float16(test_mat_T[idy,idx]).view('H'))[2:].zfill(16)
            if idx == 0:
                data = "data_a <= 16'b%s;\n" % bin(np.float16(test_mat_T[idy,idx]).view('H'))[2:].zfill(16)
                f.write(data)
            elif idx == 1:
                data = "data_b <= 16'b%s;\n" % bin(np.float16(test_mat_T[idy, idx]).view('H'))[2:].zfill(16)
                f.write(data)
            else:
                data = "data_c <= 16'b%s;\n" % bin(np.float16(test_mat_T[idy, idx]).view('H'))[2:].zfill(16)
                f.write(data)
                f.write("#10\n")

    f.close()
    wb.save('test3.xlsx')

